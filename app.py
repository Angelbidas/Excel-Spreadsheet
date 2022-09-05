"""print('Angel Bidas')
print('o----')
print(' ||||')
print('*' * 10)
name = 'John Smith'
age = '20'
is_New = True
name1 = input('What is your name? ')
color = input('What is your fav color? ')
print(name1 + ' Likes ' + color)
birth_year = input('Birth Year: ')
age1 = 2022-int(birth_year)
print(age1)
print(type(age1))
course = 'Python for Beginners'
print(course[0:3])
name2 = 'Jennifer'
print(name2[1:-1])
first = 'John'
last = 'Smith'
message = first + ' [' + last + ' ] is a coder'
message1 = f'{first} [{last}] is a coder'
print(message)
print(message1)
course1 = 'Python for Beginners'
print(len(course1))
print(course1.upper())
print(course1.lower())
print(course1.find('P'))
print(course1.find('p'))
print(course1.replace('Python', 'Python for Beginners'))
print('Python' in course1)
print('python' in course1)
print(course1.title())
print(10/3)
print(10//3)
print(10*3)
print(10**3)
print(abs(-2.9))
print(round(2.9))
import math
print(math.floor(2.9))
print(math.ceil(2.9))
print(math.sin(30))
is_hot = True
is_cold = False
if is_hot:
    print("its a hot day")
    print("Drink plenty of water")
elif is_cold:
    print("its a cold day")
    print("wear a warm clothes")
else:
    print("its a lovely day")
print("have a good day")
price = 1000000
has_good_credit = True

if has_good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.2 * price
print(f"down payment: $ {down_payment}")
weight = int(input("Weight:"))
unit = input("(L)bs or (K)g")

if unit.upper() == 'L':
    converted = weight*4.15
    print(f"You are {converted} kilos")
else:
    converted = weight/4.15
    print(f"you are {converted} pounds")
secret_number = 9
guess_count = 0
guess_limit = 3
while guess_count < guess_limit:
     guess = int(input('Guess:'))
     guess_count += 1
     if secret_number == guess:
         print("You Won!")
         break
else:
    print("You Failed!")
command = ""
started = False
while True:
    command = input(">").lower()
    if command == "Start":
        if started:
         print("Car is already started")
        else:
            started:True
            print("Car is started")
    elif command == "Stop":
        if not started:
            print("Car stopped")
        else:
            started = False
            print("Car is already stopped")
    elif command == "Help":
         print("Start-to Start the car")
         print("Stop-to Stopped the car")
         print("Quit-to quit")
    elif command == "Quit":
        break
    else:
        print("Sorry,I don't understand that")
for item in range(5, 10, 2):
    print(item)

prices = [10, 20, 30]
total = 0
for price in prices:
    total += price
print(f"Total:{total}")
# nested loops
for x in range(4):
    for y in range(3):
        print(f'({x}, {y})')
numbers = [5, 2, 5, 2, 2]
for x_count in numbers:
    print('x' * x_count)
# lists
names = ['John', 'Bob', 'Mosh', 'Sarah']
print(names[2:4])
names[0] = 'Jon'
print(names[2:])
print(names)
# largest number in list
numbers = [3, 6, 2, 8, 4, 10]
max = numbers[0]
for number in numbers:
    if number > max:
        max = number
print(max)
# 2D dimension
matrix = [
     [1, 2, 3],
     [4, 5, 6],
     [7, 8, 9]
 ]
matrix[0][1] = 20
print(matrix[0][1])
for row in matrix:
    for item in row:
        print(item)
numbers = [5, 2, 1, 7, 4]
numbers.append(20)
numbers.insert(0, 10)
numbers.remove(2)
print(numbers)
numbers.sort()
numbers.reverse()
print(numbers)
numbers2 = numbers.copy()
numbers.append(10)
print(numbers2)
# remove duplicate in lists
from typing import Dict

numbers = [2, 2, 4, 6, 3, 4, 6, 1]
uniques = []
for number in numbers:
    if number not in uniques:
        uniques.append(number)
print(uniques)
# packing in tuples
coordinates = (1, 2, 3)
coordinates[0] * coordinates[1] * coordinates[2]
x = coordinates[0]
y = coordinates[1]
z = coordinates[2]
x, y, z = coordinates
print(coordinates)
# dictionaries
customer = {
    "name": "John Smith",
    "age":30,
    "is_verified": True
}
print(customer["name"])
customer["birthdate"] = "jan 1999"
print(customer["birthdate"])
# ex2
phone = input("Phone: ")
digits_mapping = {
    "1": "One",
    "2": "Two",
    "3": "Three",
    "4": "Four"
}
output = ""
for ch in phone:
    output += digits_mapping.get(ch, "!") + " "
print(output)
# emoji converter


def emoji_converter(message):
    words = message.split(' ')
    emojis = {
        ":)": "😀",
        ":(": "😌"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
        return output


message = input(">")
print(emoji_converter(message))
# function ex2


def greet_user(first_name, second_name):
    print(f'Hi {first_name} {second_name} ')
    print('Welcome aboard')


print('start')
greet_user('John', 'smith')
print('Finish')
# return value


def square(number):
    return number * number


result = square(3)
print(result)
# exception
try:
    age = int(input('Age:'))
    income = 2000
    risk = income/age
    print(age)
except ZeroDivisionError:
    print('Age cannot be 0')
except ValueError:
    print('Invalid value')

# Classes
# we capitalise the first letter of every word


class Point:
    def __init__(self, x, y):   # constructor
        self.x = x
        self.y = y
    def move(self):
        print("move")

    def draw(self):
        print("draw")


point = Point(10, 20)
point.x = 11
print(point.x)
point.draw()
# ex2

class Person:
    def __init__(self,name):
        self.name = name

    def talk(self):
        print(f'Hi,I am {self.name}')


ange = Person("Angel Bidas")
ange.talk()

bob = Person("Bob Smith")
bob.talk()
# Inheritance

class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    def bark(self):
        print("bark")


class Cat(Mammal):
    def be_annoying(self):
        print("annoying")


dog1 = Dog()
dog1.bark()

cat1 = Cat()
cat1.be_annoying()
import converters
from converters import kg_to_lbs

kg_to_lbs(100)

print(converters.kg_to_lbs(70))

from utils import find_max


numbers = [10, 3, 6, 2]
maximum = find_max(numbers)
print(maximum)
# package
from ecommerce.shipping import calc_shipping

calc_shipping()
# built_in library
import random

for i in range(3):
    print(random.randint(10,20))
# Names ex
members = ['john', 'Mary', 'Bob', 'Mosh']
leader = random.choice(members)
print(leader)
# dice

import random


class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())
# file and directories
from pathlib import Path

path = Path("ecommerce")
print(path.exists())
# show file
from pathlib import Path

path = Path()
for file in path.glob('*.py'):
    print(file)

# show all files and directories

from pathlib import Path

path = Path()
for file in path.glob('*'):
    print(file)"""

import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
              min_row=2,
              max_row=sheet.max_row,
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)






































